#### 1. 3611에서 S90803에 해당하는 CC NUM을 엑셀파일로 추출한다.
from collections import defaultdict
import time
from openpyxl.workbook.workbook import Workbook
from pandas.core.base import DataError
from pandas.core.frame import DataFrame
import pygetwindow as gw
import pyautogui
import datetime
from pywinauto import Application
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import islice
import pandas as pd
import numpy as np
import xlrd
import win32com.client as win32

import os
import sys
import pywinauto

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', -1)

sys.stdout = open('terminal_peprdown.txt', 'w',encoding='UTF-8')

today = datetime.datetime.today() # 현재 date 추출하고 today변수에 저장

print(today)
print(today.month)

if today.month == 1: 
    sapyear = today.year
    sapmth = 12
else:
    sapyear = today.year + 1
    sapmth = today.month - 1 
today_str = today.strftime('%Y_%m_%d_%H%M')






##
##
##
####
###################################
###
##
##
## def형식으로 변환하기
def PEPR_3611(S_ACC, CY_CC, SECON_ACC, EXCLU_CYCC, EXCLU_CC): #EXCLU_CYCC:CY-399|IC-999, EXCLU_CC:3744
    
    os.system("PEPR_3611_DOWN_V2.vbs {} {} {}".format(sapyear, sapmth, S_ACC)) #Arg 0,1,2

    procs = pywinauto.findwindows.find_elements()

    for proc in procs:
        if proc.name == 'Basis (1)의 워크시트 - Excel':
            break # 해당 title 가진 파일이 있을 때 for문 탈출
        
        
    app = Application(backend='uia').connect(title="Basis (1)의 워크시트 - Excel")
    dig = app['Basis (1)의 워크시트 - Excel']


    dig.child_window(title="다른 이름으로 저장", control_type="Button").click()
    today_str = today.strftime('%Y_%m_%d_%H%M')
    dig.child_window(title="파일 이름:", auto_id="1001", control_type="Edit").type_keys('{}_3611_{}{}_{}.xlsx'.format(S_ACC,sapyear,sapmth,today_str))
    dig.child_window(title="저장(S)", auto_id="1", control_type="Button").click()
    dig.child_window(title="닫기", control_type="Button", found_index=0).click()
    
    # 그 다음 이 파일을 pandas로 체크하여 cy-399xx가 있는지 확인하고 그 값이 있다면 이를 변수에 저장해서
    df = pd.read_excel('{}_3611_{}{}_{}.xlsx'.format(S_ACC,sapyear,sapmth,today_str), sheet_name='Sheet1', engine='openpyxl')

    print(df)
    df = df.drop_duplicates(['Cost Ctr'], keep = 'first', ignore_index = True)
    df = df['Cost Ctr']
    print(df)
    # cy399list = ['CY-399CZ', 'CY-399MT', 'CY-399MW', 'CY-399SD', 'CY-399VF', 'CY-399VJ']
    # df_cy399 = df[df.isin(cy399list)]
    df_cy = df[df.str.contains(CY_CC)]

    print(df_cy)
    df_cy  = df_cy.head(1)
    df_cy = df_cy.to_clipboard(index=False, header=False)
    print(df_cy)


    ## 해당 변수를 클립보드에 넣었음, vbs파일을 돌려서 sap에서 그 값을 찾도록 한다.
    os.system("3611_cy399_down_v5.vbs {} {} {} {}".format(sapyear, sapmth, CY_CC , today_str)) 
    
    ## xls로 export된 것을 xlsx형식으로 변환
    fname = "{}_{}{}_exd{}.xls".format(CY_CC, sapyear,sapmth, today_str) 
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)

    wb.SaveAs(fname+"x", FileFormat = 51) # file format = 51 is xlsx format
    wb.Close() # Fileformat = 56 is xls extension
    excel.Application.Quit()

    ## xlsx를 df형식으로 불러와 데이터처리
    df2 = pd.read_excel('{}_{}{}_exd{}.xlsx'.format(CY_CC, sapyear, sapmth, today_str),usecols="D,I",skiprows=6, header=None , engine="openpyxl") # CC와GL정보 있는 COLS만 가져오고, 빈값은 날림.
    df2.columns = "Costcenter", "gl"
    print(df2)

    ## s87410만 불러옴 SECON_ACC에 S87410을 넣으면 될듯 #함수화 하는데 어디서 끊어야할지 모르곘음...
    condition = (df2.gl == SECON_ACC)
    df2 = df2[condition]
    print(df2)

    ## s87410을 가지는 cc중 중복되는 값들은 제거한다.
    df2 = df2.drop_duplicates(['Costcenter'], keep = 'first', ignore_index = True)
    df2 = df2['Costcenter']
    print(df2)
    print("df2를 프린트하였습니다.")

    df2 = df2[~df2.str.contains(EXCLU_CYCC, na=False)] # cy를 포함하지 않는것만 불러온다.
    df2_cy = df2[~df2.isin({EXCLU_CC})]
    print(df2_cy)
    print("df2_cy를 프린트하였습니다.")
    df_wocy = df[~df.str.contains(EXCLU_CYCC, na=False)] #cy399sd도 포함해야하는 달이있음 어떻게 이를 코드로 구현할까?
    df_wocy = df_wocy[~df_wocy.isin({EXCLU_CC})]
    # df_cy399sd = pd.DataFrame({'CY-399SD'}) # 그냥 값체크없이 CY-399SD를 하나의 DF로 만들어서 추가함 # 이 라인은 S90820에만 필요함
    # result1 = pd.concat([df_wocy,df2_cy399,df_cy399sd])
    global result1
    result1 = pd.concat([df_wocy,df2_cy], ignore_index=True) #안에 []를 ()으로 고쳤음
    
    
    print(result1)
    print("result1을 프린트하였습니다.")
    df3 = pd.DataFrame(result1, columns=['Costcenter'])#######################이게 cctr중복제거하려고 넣은값 만약에 오류해결안되면 지우자#############
    df3 = df3.astype('str') # 중복제거를 위해서 str형식을 변환시킴
    print(df3)
    print("df3를 프린트하였습니다.")
    df3 = df3.drop_duplicates(['Costcenter'], keep = 'first', ignore_index = True, inplace=False) #######################이게 cctr중복제거하려고 넣은값 만약에 오류해결안되면 지우자############################################
    print(df3)
    print("df3를 프린트하였습니다.")
    
    result1 = pd.DataFrame(df3)
    print(result1)
    print("result1을 프린트하였습니다.")
    
    s_ACC_cc = result1.to_clipboard(index=False, header=False)
    print(s_ACC_cc)
    return result1
    
    
    ## 그 다음 cy399 3611파일에서 필요한 cost center넘버를 다시 가지고 와서 처음에 뽑은 cc num + cy399 cc num
    ## 이 두개를 동시에 sap 3613에 넣고 돌린다.
    ## 아래는 S90810대상으로 3613뽑는 코드를 S90820에 맞게 변환한 것이다. 혹시나 에러가 나는지 보고 트러블슈팅하자.





    # ### 6. 해당 s계정의 cost center를 클립보드에 받고 이를 SAP3613 돌려서 엑셀 XLS파일로 내려받음
def PEPR_3613(S_ACC):
    os.system("PEPR_3613_DOWN_V1.vbs {} {} {} {}".format(sapyear, sapmth, S_ACC, today_str)) #Arg 0,1,2,3



    # ### 7. XLS파일을 xlsx형식으로 변환 pywinauto를 안쓰고 변환할 수 는 없을까? win32 사용
    fname = "{}_3613_{}{}_exd{}.xls".format(S_ACC ,sapyear, sapmth , today_str)
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)

    wb.SaveAs(fname+"x", FileFormat = 51) # fileformat 51이 xlsx뜻함
    wb.Close()
    excel.Application.Quit()

    ## 8. openpyxl로 3613에서 추출한 엑셀 편집

    wb = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format(S_ACC ,sapyear, sapmth , today_str))
    ws = wb.active
    ws.title = "%s_3613" % S_ACC
    cpws = wb.copy_worksheet(ws)
    cpws.title = "Copied sheet"


    cpws.delete_rows(1, 33)
    cpws.delete_rows(2, 2)
    cpws.delete_cols(6, 100)

    # 셀의 특정값이 있다면 행 아래를 지우고 싶었는데 안됨...
    column_max = cpws.max_column
    row_max = cpws.max_row
    for col_num in range(1, column_max+1):
        for row_num in range(1, row_max+1):
            if str(cpws.cell(row = row_num, column= col_num).value) == "** Total":
                cpws.delete_rows(row_num, 1000)
            
                


    wb.save("{}_3613_{}{}_exd{}.xlsx".format(S_ACC ,sapyear, sapmth , today_str))
    wb.close()

    wb2 = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format(S_ACC ,sapyear, sapmth , today_str))
    ws2 = wb2['Copied sheet']
    ws2.cell(row=1, column=2).value = 'cc'
    ws2.cell(row=1, column=5).value = 'actcost'
    df2 = pd.DataFrame(ws2.values)

    data2 = ws2.values
    cols2 = next(data2)[1:]
    data2 = list(data2)
    idx2 = [r[0] for r in data2]
    data2 = (islice(r, 1, None) for r in data2)

    df2 = pd.DataFrame(data2, columns=cols2)

    # print(df2)

    df2[['star','blank', 'conum', 'coname']] = pd.DataFrame(df2.cc.str.split(' ', 3).tolist())
    df3 = df2[['star', 'conum', 'coname', 'actcost']]

    print(df3)

    df3.to_excel("{}_3613_{}{}_exd{}.xlsx".format(S_ACC ,sapyear, sapmth , today_str), index=False)

    wb3 = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format(S_ACC,sapyear, sapmth , today_str))
    ws3 = wb3.active
    ws3.insert_cols(1)

    for r in ws3.rows:
        row_index = r[0].row
        print(r[0])
        ws3['A'+str(row_index)] = '=IF(B'+str(row_index)+'="",A'+str(row_index+1)+',C'+str(row_index)+')'


    wb3.save("{}_3613_{}{}_exd{}.xlsx".format(S_ACC,sapyear, sapmth , today_str))
    return


def PEPR_3613_excel_edit(S_ACC):
    
    fname = "{}_3613_{}{}_exd{}.xlsx".format(S_ACC ,sapyear, sapmth , today_str)
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)
    os.system("excel_valuecopy_v1.vbs")
    wb.SaveAs("{}_3613_{}{}_exd{}_V2.xlsx".format(S_ACC ,sapyear, sapmth , today_str), FileFormat = 51) # fileformat 51이 xlsx뜻함
    wb.Close()
    excel.Application.Quit()
    # wb4 = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format(S_ACC ,sapyear, sapmth , today_str))
    # ws4 = wb4.active
   
    # df4 = pd.DataFrame(ws4.values)

    # print(df4)

    # df4.to_excel("{}_3613_valuecopy_{}{}_exd{}.xlsx".format(S_ACC ,sapyear, sapmth , today_str), index=False)

    # wb3 = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format(S_ACC,sapyear, sapmth , today_str))
    # ws3 = wb3.active
    # ws3.insert_cols(1)

    # for r in ws3.rows:
    #     row_index = r[0].row
    #     print(r[0])
    #     ws3['A'+str(row_index)] = '=IF(B'+str(row_index)+'="",A'+str(row_index+1)+',C'+str(row_index)+')'


    # wb3.save("{}_3613_{}{}_exd{}.xlsx".format(S_ACC,sapyear, sapmth , today_str))
    return


     
##
##
####
############################################## 함수형식으로 바꾸어서 테스트하기
##
#def PEPR_3611(S_ACC, CY_CC, SECON_ACC, EXCLU_CYCC, EXCLU_CC): #S_ACC: 'S90820', CY_CC:'CY-399', SECON_ACC:'S87410', EXCLU_CYCC:'CY-399|IC-999', EXCLU_CC:3744

PEPR_3611('S90820','CY-399', 'S87410', 'CY-399|IC-999', 3744)
df_cy399sd = pd.DataFrame({'CY-399SD'}, columns=['Costcenter']) # 그냥 값체크없이 CY-399SD를 하나의 DF로 만들어서 추가함 # 이 라인은 S90820에만 필요함
result1 = pd.concat([result1 ,df_cy399sd], ignore_index=True)

print(result1)
df4 = pd.DataFrame(result1, columns=['Costcenter'])
df4 = df4.astype('str')
df4 = df4.drop_duplicates(['Costcenter'], keep = 'first', ignore_index = True, inplace=False) #######################이게 cctr중복제거하려고 넣은값 만약에 오류해결안되면 지우자######################

s_ACC_cc = df4.to_clipboard(index=False, header=False)
print(s_ACC_cc)

PEPR_3613('S90820') # 같은 코스트센터가 두번 중복으로 들어가서 값이 더블되는 오류있음 cctr중복체크후 3613돌리는 로직 필요
PEPR_3613_excel_edit('S90820')  # 값복사하는코드 안되면 지우자.



# PEPR_3611('S90801', 'CY-499', 'S87420', 'CY-499', 99999)
# PEPR_3613('S90801') 
# # S90801 MGK는 MGK,VK 나누는 코드가 필요, 함수로 만들어보자.
# df = pd.read_excel("{}_3613_{}{}_exd{}.xlsx".format('S90801',sapyear, sapmth , today_str), sheet_name='Sheet1', engine='openpyxl',index_col=0)
# print(df)
# index_star = df['star'] = "*".index
# print(index_star)
# df.drop(index_star, axis=0)
# print(df)
###################################################


### 3611에서 각s계정의 cost center엑셀로 추출
# os.system("PEPR_3611_DOWN_V2.vbs {} {} {}".format(sapyear, sapmth, "S90803")) #Arg 0,1,2


# procs = pywinauto.findwindows.find_elements()

# for proc in procs:
#     if proc.name == 'Basis (1)의 워크시트 - Excel':
#         break # 해당 title 가진 파일이 있을 때 for문 탈출
    
    
# app = Application(backend='uia').connect(title="Basis (1)의 워크시트 - Excel")
# dig = app['Basis (1)의 워크시트 - Excel']


# dig.child_window(title="다른 이름으로 저장", control_type="Button").click()
# today_str = today.strftime('%Y_%m_%d_%H%M')
# dig.child_window(title="파일 이름:", auto_id="1001", control_type="Edit").type_keys('S90803_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str))
# dig.child_window(title="저장(S)", auto_id="1", control_type="Button").click()
# dig.child_window(title="닫기", control_type="Button", found_index=0).click()
# # # dig.print_control_identifiers()



# # # #### 2. 3611에서 S90804 CC NUM 추출
# os.system("PEPR_3611_DOWN_V2.vbs {} {} {}".format(sapyear, sapmth, "S90804")) #Arg 0,1,2

# procs = pywinauto.findwindows.find_elements()

# for proc in procs:
#     if proc.name == 'Basis (1)의 워크시트 - Excel':
#         break # 해당 title 가진 파일이 있을 때 for문 탈출
    
    
# app = Application(backend='uia').connect(title="Basis (1)의 워크시트 - Excel")
# dig = app['Basis (1)의 워크시트 - Excel']


# dig.child_window(title="다른 이름으로 저장", control_type="Button").click()
# today_str = today.strftime('%Y_%m_%d_%H%M')
# dig.child_window(title="파일 이름:", auto_id="1001", control_type="Edit").type_keys('S90804_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str))
# dig.child_window(title="저장(S)", auto_id="1", control_type="Button").click()
# dig.child_window(title="닫기", control_type="Button", found_index=0).click()


































# # # # #### 3. 3611에서 S90820 CC NUM 추출
# os.system("PEPR_3611_DOWN_V2.vbs {} {} {}".format(sapyear, sapmth, "S90820")) #Arg 0,1,2

# procs = pywinauto.findwindows.find_elements()

# for proc in procs:
#     if proc.name == 'Basis (1)의 워크시트 - Excel':
#         break # 해당 title 가진 파일이 있을 때 for문 탈출
    
    
# app = Application(backend='uia').connect(title="Basis (1)의 워크시트 - Excel")
# dig = app['Basis (1)의 워크시트 - Excel']


# dig.child_window(title="다른 이름으로 저장", control_type="Button").click()
# today_str = today.strftime('%Y_%m_%d_%H%M')
# dig.child_window(title="파일 이름:", auto_id="1001", control_type="Edit").type_keys('S90820_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str))
# dig.child_window(title="저장(S)", auto_id="1", control_type="Button").click()
# dig.child_window(title="닫기", control_type="Button", found_index=0).click()


# # 그 다음 이 파일을 pandas로 체크하여 cy-399xx가 있는지 확인하고 그 값이 있다면 이를 변수에 저장해서
# df = pd.read_excel('S90820_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str), sheet_name='Sheet1', engine='openpyxl')

# print(df)
# df = df.drop_duplicates(['Cost Ctr'], keep = 'first', ignore_index = True)
# df = df['Cost Ctr']
# print(df)
# # cy399list = ['CY-399CZ', 'CY-399MT', 'CY-399MW', 'CY-399SD', 'CY-399VF', 'CY-399VJ']
# # df_cy399 = df[df.isin(cy399list)]
# df_cy399 = df[df.str.contains('CY-399')]

# print(df_cy399)
# df_cy399  = df_cy399.head(1)
# df_cy399 = df_cy399.to_clipboard(index=False, header=False)
# print(df_cy399)


# ## 해당 변수를 클립보드에 넣었음, vbs파일을 돌려서 sap에서 그 값을 찾도록 한다.
# os.system("3611_cy399_down_v5.vbs {} {} {} {}".format(sapyear, sapmth, "cy399", today_str)) 
 
# ## xls로 export된 것을 xlsx형식으로 변환
# fname = "{}_{}{}_exd{}.xls".format("cy399", sapyear,sapmth, today_str) 
# excel = win32.gencache.EnsureDispatch('Excel.Application')
# wb = excel.Workbooks.Open(fname)

# wb.SaveAs(fname+"x", FileFormat = 51) # file format = 51 is xlsx format
# wb.Close() # Fileformat = 56 is xls extension
# excel.Application.Quit()

# ## xlsx를 df형식으로 불러와 데이터처리
# df2 = pd.read_excel('{}_{}{}_exd{}.xlsx'.format("cy399", sapyear, sapmth, today_str),usecols="D,I",skiprows=6, header=None , engine="openpyxl") # CC와GL정보 있는 COLS만 가져오고, 빈값은 날림.
# df2.columns = "Costcenter", "gl"
# print(df2)

# ## s87410만 불러옴
# condition = (df2.gl =='S87410')
# df2 = df2[condition]
# print(df2)

# ## s87410을 가지는 cc중 중복되는 값들은 제거한다.
# df2 = df2.drop_duplicates(['Costcenter'], keep = 'first', ignore_index = True)
# df2 = df2['Costcenter']
# print(df2)

# df2 = df2[~df2.str.contains('CY-399|IC-999', na=False)] # cy를 포함하지 않는것만 불러온다.
# df2_cy399 = df2[~df2.isin({3744})]
# print(df2_cy399)
# df_wocy = df[~df.str.contains('CY-399|IC-999', na=False)] #cy399sd도 포함해야하는 달이있음 어떻게 이를 코드로 구현할까?
# df_wocy = df_wocy[~df_wocy.isin({3744})]
# df_cy399sd = pd.DataFrame({'CY-399SD'}) # 그냥 값체크없이 CY-399SD를 하나의 DF로 만들어서 추가함 # 왜 cy-399sd가 아닌 #cy-399sd가 들어갈까?
# result1 = pd.concat([df_wocy,df2_cy399,df_cy399sd])
# print(result1)

# s90820cc = result1.to_clipboard(index=False, header=False)
# print(s90820cc)

# ## 그 다음 cy399 3611파일에서 필요한 cost center넘버를 다시 가지고 와서 처음에 뽑은 cc num + cy399 cc num
# ## 이 두개를 동시에 sap 3613에 넣고 돌린다.
# ## 아래는 S90810대상으로 3613뽑는 코드를 S90820에 맞게 변환한 것이다. 혹시나 에러가 나는지 보고 트러블슈팅하자.

# # ### 6. 해당 s계정의 cost center를 클립보드에 받고 이를 SAP3613 돌려서 엑셀 XLS파일로 내려받음
# os.system("PEPR_3613_DOWN_V1.vbs {} {} {} {}".format(sapyear, sapmth, "S90820", today_str)) #Arg 0,1,2,3



# # ### 7. XLS파일을 xlsx형식으로 변환 pywinauto를 안쓰고 변환할 수 는 없을까? win32 사용
# fname = "{}_3613_{}{}_exd{}.xls".format("S90820",sapyear, sapmth , today_str)
# excel = win32.gencache.EnsureDispatch('Excel.Application')
# wb = excel.Workbooks.Open(fname)

# wb.SaveAs(fname+"x", FileFormat = 51) # fileformat 51이 xlsx뜻함
# wb.Close()
# excel.Application.Quit()

# ## 8. openpyxl로 3613에서 추출한 엑셀 편집

# wb = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format("S90820",sapyear, sapmth , today_str))
# ws = wb.active
# ws.title = "S90820_3613"
# cpws = wb.copy_worksheet(ws)
# cpws.title = "Copied sheet"


# cpws.delete_rows(1, 33)
# cpws.delete_rows(2, 2)
# cpws.delete_cols(6, 100)

# # 셀의 특정값이 있다면 행 아래를 지우고 싶었는데 안됨...
# column_max = cpws.max_column
# row_max = cpws.max_row
# for col_num in range(1, column_max+1):
#     for row_num in range(1, row_max+1):
#         if str(cpws.cell(row = row_num, column= col_num).value) == "** Total":
#             cpws.delete_rows(row_num, 1000)
          
            


# wb.save("{}_3613_{}{}_exd{}.xlsx".format("S90820",sapyear, sapmth , today_str))
# wb.close()

# wb2 = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format("S90820",sapyear, sapmth , today_str))
# ws2 = wb2['Copied sheet']
# ws2.cell(row=1, column=2).value = 'cc'
# ws2.cell(row=1, column=5).value = 'actcost'
# df2 = pd.DataFrame(ws2.values)

# data2 = ws2.values
# cols2 = next(data2)[1:]
# data2 = list(data2)
# idx2 = [r[0] for r in data2]
# data2 = (islice(r, 1, None) for r in data2)

# df2 = pd.DataFrame(data2, columns=cols2)

# # print(df2)

# df2[['star','blank', 'conum', 'coname']] = pd.DataFrame(df2.cc.str.split(' ', 3).tolist())
# df3 = df2[['star', 'conum', 'coname', 'actcost']]

# print(df3)

# df3.to_excel("{}_3613_{}{}_exd{}.xlsx".format("S90820",sapyear, sapmth , today_str), index=False)

# wb3 = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format("S90820",sapyear, sapmth , today_str))
# ws3 = wb3.active
# ws3.insert_cols(1)

# for r in ws3.rows:
#     row_index = r[0].row
#     print(r[0])
#     ws3['A'+str(row_index)] = '=IF(B'+str(row_index)+'="",A'+str(row_index+1)+',C'+str(row_index)+')'


# wb3.save("{}_3613_{}{}_exd{}.xlsx".format("S90820",sapyear, sapmth , today_str))































































# # # #### 4. 3611에서 S90801 CC NUM 추출
# os.system("PEPR_3611_DOWN_V2.vbs {} {} {}".format(sapyear, sapmth, "S90801")) #Arg 0,1,2

# procs = pywinauto.findwindows.find_elements()

# for proc in procs:
#     if proc.name == 'Basis (1)의 워크시트 - Excel':
#         break # 해당 title 가진 파일이 있을 때 for문 탈출
    
    
# app = Application(backend='uia').connect(title="Basis (1)의 워크시트 - Excel")
# dig = app['Basis (1)의 워크시트 - Excel']


# dig.child_window(title="다른 이름으로 저장", control_type="Button").click()
# today_str = today.strftime('%Y_%m_%d_%H%M')
# dig.child_window(title="파일 이름:", auto_id="1001", control_type="Edit").type_keys('S90801_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str))
# dig.child_window(title="저장(S)", auto_id="1", control_type="Button").click()
# dig.child_window(title="닫기", control_type="Button", found_index=0).click()


## 5. 3611에서 S90805 추출
# os.system("PEPR_3611_DOWN_V2.vbs {} {} {}".format(sapyear, sapmth, "S90805")) #Arg 0,1,2

# procs = pywinauto.findwindows.find_elements()

# for proc in procs:
#     if proc.name == 'Basis (1)의 워크시트 - Excel':
#         break # 해당 title 가진 파일이 있을 때 for문 탈출
    
    
# app = Application(backend='uia').connect(title="Basis (1)의 워크시트 - Excel")
# dig = app['Basis (1)의 워크시트 - Excel']


# dig.child_window(title="다른 이름으로 저장", control_type="Button").click()
# today_str = today.strftime('%Y_%m_%d_%H%M')
# dig.child_window(title="파일 이름:", auto_id="1001", control_type="Edit").type_keys('S90805_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str))
# dig.child_window(title="저장(S)", auto_id="1", control_type="Button").click()
# dig.child_window(title="닫기", control_type="Button", found_index=0).click()

# os.system("PEPR_3611_DOWN_V2.vbs {} {} {}".format(sapyear, sapmth, "S90808")) #Arg 0,1,2

# procs = pywinauto.findwindows.find_elements()

# for proc in procs:
#     if proc.name == 'Basis (1)의 워크시트 - Excel':
#         break # 해당 title 가진 파일이 있을 때 for문 탈출
    
    
# app = Application(backend='uia').connect(title="Basis (1)의 워크시트 - Excel")
# dig = app['Basis (1)의 워크시트 - Excel']


# dig.child_window(title="다른 이름으로 저장", control_type="Button").click()
# today_str = today.strftime('%Y_%m_%d_%H%M')
# dig.child_window(title="파일 이름:", auto_id="1001", control_type="Edit").type_keys('S90808_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str))
# dig.child_window(title="저장(S)", auto_id="1", control_type="Button").click()
# dig.child_window(title="닫기", control_type="Button", found_index=0).click()


## 5. 3611에서 S90802 추출
# os.system("PEPR_3611_DOWN_V2.vbs {} {} {}".format(sapyear, sapmth, "S90802")) #Arg 0,1,2

# procs = pywinauto.findwindows.find_elements()

# for proc in procs:
#     if proc.name == 'Basis (1)의 워크시트 - Excel':
#         break # 해당 title 가진 파일이 있을 때 for문 탈출
    
    
# app = Application(backend='uia').connect(title="Basis (1)의 워크시트 - Excel")
# dig = app['Basis (1)의 워크시트 - Excel']


# dig.child_window(title="다른 이름으로 저장", control_type="Button").click()
# today_str = today.strftime('%Y_%m_%d_%H%M')
# dig.child_window(title="파일 이름:", auto_id="1001", control_type="Edit").type_keys('S90802_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str))
# dig.child_window(title="저장(S)", auto_id="1", control_type="Button").click()
# dig.child_window(title="닫기", control_type="Button", found_index=0).click()










#######
#######
#######














# # 7. 3611에서 S90810 추출
# os.system("PEPR_3611_DOWN_V2.vbs {} {} {}".format(sapyear, sapmth, "S90810")) #Arg 0,1,2

# procs = pywinauto.findwindows.find_elements()

# for proc in procs:
#     if proc.name == 'Basis (1)의 워크시트 - Excel':
#         break # 해당 title 가진 파일이 있을 때 for문 탈출
    
    
# app = Application(backend='uia').connect(title="Basis (1)의 워크시트 - Excel")
# dig = app['Basis (1)의 워크시트 - Excel']


# dig.child_window(title="다른 이름으로 저장", control_type="Button").click()
# today_str = today.strftime('%Y_%m_%d_%H%M')
# dig.child_window(title="파일 이름:", auto_id="1001", control_type="Edit").type_keys('S90810_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str))
# dig.child_window(title="저장(S)", auto_id="1", control_type="Button").click()
# dig.child_window(title="닫기", control_type="Button", found_index=0).click()
# time.sleep(2)

# # ### 5. 엑셀파일을 열어 cost center의 중복값을 제거
# # wb = openpyxl.load_workbook('S90810_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str))
# # ws = wb.active

# # # 모든 값을 DATA FRAME에 포함하고 INDEX와 COL 이름은 0,1,2 로 사용
# # df = DataFrame(ws.values)

# # data = ws.values
# # cols = next(data)[1:]
# # data = list(data)
# # idx = [r[0] for r in data]
# # data = (islice(r, 1, None) for r in data)
# # # 첫 행이 index이고 첫열이 column명일 경우 아래와 같이 사용
# # df = DataFrame(data, columns=cols)


# ##
# df = pd.read_excel('S90810_3611_{}{}_{}.xlsx'.format(sapyear,sapmth,today_str), sheet_name='Sheet1', engine='openpyxl')

# print(df)
# df = df.drop_duplicates(['Cost Ctr'], keep = 'first', ignore_index = True)
# df = df['Cost Ctr']
# print(df)
# # cy399list = ['CY-399CZ', 'CY-399MT', 'CY-399MW', 'CY-399SD', 'CY-399VF', 'CY-399VJ']
# # df_cy399 = df[df.isin(cy399list)]



# # 중복값 제거, 어떤 열을 남길지 keep으로 선택, ignore_index로 사라진 행을 날려버림.

# df_ccnum = df.to_clipboard(index=False, header=False) # col header를 날리기 위해서 false넣음
# print(df_ccnum)

# ### 6. 해당 s계정의 cost center를 클립보드에 받고 이를 SAP3613 돌려서 엑셀 XLS파일로 내려받음
# os.system("PEPR_3613_DOWN_V1.vbs {} {} {} {}".format(sapyear, sapmth, "S90810", today_str)) #Arg 0,1,2,3



# # ### 7. XLS파일을 xlsx형식으로 변환 pywinauto를 안쓰고 변환할 수 는 없을까? win32 사용

# fname = "{}_3613_{}{}_exd{}.xls".format("S90810",sapyear, sapmth , today_str)
# excel = win32.gencache.EnsureDispatch('Excel.Application')
# wb = excel.Workbooks.Open(fname)

# wb.SaveAs(fname+"x", FileFormat = 51) # fileformat 51이 xlsx뜻함
# wb.Close()
# excel.Application.Quit()

# ## 8. openpyxl로 3613에서 추출한 엑셀 편집

# wb = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format("S90810",sapyear, sapmth , today_str))
# ws = wb.active
# ws.title = "S90810_3613"
# cpws = wb.copy_worksheet(ws)
# cpws.title = "Copied sheet"


# cpws.delete_rows(1, 33)
# cpws.delete_rows(2, 2)
# cpws.delete_cols(6, 100)

# # 셀의 특정값이 있다면 행 아래를 지우고 싶었는데 안됨...
# column_max = cpws.max_column
# row_max = cpws.max_row
# for col_num in range(1, column_max+1):
#     for row_num in range(1, row_max+1):
#         if str(cpws.cell(row = row_num, column= col_num).value) == "** Total":
#             cpws.delete_rows(row_num, 1000)
          
            


# wb.save("{}_3613_{}{}_exd{}.xlsx".format("S90810",sapyear, sapmth , today_str))
# wb.close()

# wb4 = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format("S90810",sapyear, sapmth , today_str))
# ws4 = wb4['Copied sheet']
# ws4.cell(row=1, column=2).value = 'cc'
# ws4.cell(row=1, column=5).value = 'actcost'
# df4 = pd.DataFrame(ws4.values)

# data4 = ws4.values
# cols4 = next(data4)[1:]
# data4 = list(data4)
# idx4 = [r[0] for r in data4]
# data4 = (islice(r, 1, None) for r in data4)

# df4 = pd.DataFrame(data4, columns=cols4)

# # print(df2)

# df4[['star','blank', 'conum', 'coname']] = pd.DataFrame(df4.cc.str.split(' ', 3).tolist())
# df5 = df4[['star', 'conum', 'coname', 'actcost']]

# print(df5)

# df5.to_excel("{}_3613_{}{}_exd{}.xlsx".format("S90810",sapyear, sapmth , today_str), index=False)

# wb5 = openpyxl.load_workbook("{}_3613_{}{}_exd{}.xlsx".format("S90810",sapyear, sapmth , today_str))
# ws5 = wb5.active
# ws5.insert_cols(1)

# for r in ws5.rows:
#     row_index = r[0].row
#     print(r[0])
#     ws5['A'+str(row_index)] = '=IF(B'+str(row_index)+'="",A'+str(row_index+1)+',C'+str(row_index)+')'


# wb5.save("{}_3613_{}{}_exd{}.xlsx".format("S90810",sapyear, sapmth , today_str))
# wb5.close()
# ##
# ##
# ##

####11

